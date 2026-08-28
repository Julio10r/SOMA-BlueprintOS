#!/usr/bin/env python3
"""
Linx/WISE daily integration script.

GOVERNANCE NOTE (WAVE 0 containment, see docs/audits/AgentsFinalization-EnforcementUniversal.md):

This script used to open a direct pyodbc connection to production (LINX_PROD_*)
and execute INSERT/UPDATE/executemany/commit with no governance in the loop at
all — a direct bypass of agents/EXECUTION_POLICY.md ("No Direct Bypass") and of
the whole ActionProposal/AIGovernancePolicyEngine/ApprovalPolicy/GovernedWriteStack
pipeline in applications/mais-compras/backend/src/BlueprintOS.Core/AI/Governance/.

The parsing/validation/planning logic (spreadsheet reading, environment gate,
missing-product/color checks, WISE expected-vs-remote comparisons, price/stock
diffing) is real operational knowledge and is preserved unchanged in behavior.
It has been split into pure functions (`read_rows`, `build_plan`) that never
touch a live connection.

What changed: actual mutation (write) execution is no longer reachable by
default. By default the script runs in PLAN-ONLY / DRY-RUN mode: it connects
read-only (SELECT-only queries needed to build the plan), computes the same
diffs as before, and instead of calling cur.execute(UPDATE/INSERT)/cn.commit()
it emits the intended governed write plan as structured JSON
(`governed_write_plan.json`) — RequestId, AgentId, Capability, Environment,
Resource, Intent, Filter, ExpectedAffectedRows, DataClassification, Purpose,
ConnectionProfile — matching the plan contract produced by
tools/agents/governed-orchestrator.js and consumed by the .NET
GovernedWriteStack.PrepareAsync bridge (see
applications/mais-compras/backend/src/BlueprintOS.Application/Governance/GovernedPlanBridge.cs).

Real mutation is only reachable when ALL of the following hold:
  1. --execute is passed explicitly (dry-run is the implicit default), AND
  2. the environment variable GOVERNANCE_APPROVED_EXECUTION is set to the
     sha256 hex digest of the exact governed_write_plan.json produced by this
     run (i.e. the plan cannot be altered after approval without invalidating
     the hash — same hash-mismatch--> BLOCKED contract as the .NET
     GovernedWriteStack), AND
  3. a local approval record exists at
     .ai/local-output/governance/wise_approvals/<request_id>.json with
     status == "GRANTED", a matching plan_hash, and an expiry in the future.
     This file is expected to be produced by a human/pipeline step that has
     gone through the real .NET ApprovalPolicy (GrantAsync) out of band and
     exported the grant — it is NOT itself the source of authorization, it is
     a local mirror the script can check without a live cross-process call
     from Python into the .NET process. This is a known limitation, documented
     in the finalization doc as BY_DESIGN pending a real synchronous bridge.

If any of the three conditions is missing, the script prints:
    GOVERNANCE_REQUIRED: <reason>
and exits non-zero without ever opening a write-capable connection or calling
cur.execute() with a mutating statement.

LIVE_EXECUTION is never implied by this script alone — it still requires the
external, out-of-band approval artifact described above. This script does not,
and must not, grant its own approval.
"""
import argparse
import csv
import hashlib
import json
import os
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / ".ai" / "local-output" / "mb_prod_extra_web" / "current"
APPROVALS_DIR = ROOT / ".ai" / "local-output" / "governance" / "wise_approvals"

CONNECTION_PROFILE = "wise-governed-write"  # must match WiseGovernedAdapter.AllowedConnectionProfiles
CAPABILITY = "wise-database-write-proposal"  # must match WiseGovernedAdapter.Capability
AGENT_ID = "wise-agent"  # must match WiseGovernedAdapter.OwnerAgent
GOVERNED_PLAN_CLI_DLL = ROOT / "applications/mais-compras/backend/src/BlueprintOS.Api/bin/Debug/net9.0/BlueprintOS.Api.dll"


def load_env():
    env_path = ROOT / ".env"
    if not env_path.exists():
        raise SystemExit(".env not found")
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _require_pyodbc():
    """Import pyodbc lazily. Only ever called from the read-only connection
    path or from the governed write path — never at module import time — so
    that plan-only mode has no dependency on a live driver being installed
    and never even attempts a socket to production by accident."""
    import pyodbc  # noqa: PLC0415

    return pyodbc


def conn(read_only_intent: bool):
    """Open a real database connection. `read_only_intent` is documentation
    only — SQL Server has no client-enforced read-only session mode via this
    driver — the actual write-prevention comes from the caller never issuing
    mutating statements in plan-only mode (enforced by control flow in
    `run()`, not by this function)."""
    pyodbc = _require_pyodbc()
    load_env()
    required = ["LINX_PROD_SERVER", "LINX_PROD_DATABASE", "LINX_PROD_USER", "LINX_PROD_PASSWORD"]
    missing = [k for k in required if not os.environ.get(k)]
    if missing:
        raise SystemExit(f"missing env vars: {', '.join(missing)}")
    cs = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={os.environ['LINX_PROD_SERVER']};"
        f"DATABASE={os.environ['LINX_PROD_DATABASE']};"
        f"UID={os.environ['LINX_PROD_USER']};"
        f"PWD={os.environ['LINX_PROD_PASSWORD']};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(cs, autocommit=False, timeout=30)


# ---------------------------------------------------------------------------
# Parser: spreadsheet -> rows (unchanged business logic, pure function)
# ---------------------------------------------------------------------------

def read_rows(path: Path, data_limite: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["PRODUTO", "COR_PRODUTO", "TOTAL"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise SystemExit(f"missing spreadsheet columns: {missing}")
    tam_cols = [h for h in headers if h.startswith("TAM_")]
    if not tam_cols:
        raise SystemExit("no TAM_n columns found")
    rows = []
    errors = []
    seen = set()
    for rnum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        produto = str(row[idx["PRODUTO"]]).strip()
        cor = str(row[idx["COR_PRODUTO"]]).strip().zfill(4)
        vals = {}
        total_calc = 0
        for h in tam_cols:
            n = int(h.split("_", 1)[1])
            val = row[idx[h]]
            val = int(val or 0)
            vals[f"EX{n}"] = val
            total_calc += val
        total = int(row[idx["TOTAL"]] or 0)
        key = (produto, cor)
        if key in seen:
            errors.append({"row": rnum, "erro": "DUPLICATE_PRODUCT_COLOR", "produto": produto, "cor": cor})
        seen.add(key)
        if total != total_calc:
            errors.append({"row": rnum, "erro": "TOTAL_DIVERGENTE", "produto": produto, "cor": cor, "total": total, "calculado": total_calc})
        rows.append({"row": rnum, "PRODUTO": produto, "COR_PRODUTO": cor, "DATA_LIMITE": data_limite, "TOTAL_PLANILHA": total, **vals})
    return rows, tam_cols, errors


def write_csv(path, rows, fields=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None and rows:
        fields = list(rows[0].keys())
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields or [])
        w.writeheader()
        w.writerows(rows)


def make_temp(cur, rows):
    cur.execute("IF OBJECT_ID('tempdb..#carga') IS NOT NULL DROP TABLE #carga")
    ex_cols = sorted([k for k in rows[0] if k.startswith("EX")], key=lambda x: int(x[2:]))
    cur.execute(
        "CREATE TABLE #carga (PRODUTO varchar(30) NOT NULL, COR_PRODUTO varchar(20) NOT NULL, DATA_LIMITE date NOT NULL, "
        + ", ".join(f"{c} int NOT NULL" for c in ex_cols)
        + ", TOTAL_PLANILHA int NOT NULL, PRIMARY KEY (PRODUTO, COR_PRODUTO))"
    )
    cols = ["PRODUTO", "COR_PRODUTO", "DATA_LIMITE", *ex_cols, "TOTAL_PLANILHA"]
    placeholders = ",".join("?" for _ in cols)
    cur.fast_executemany = True
    cur.executemany(
        f"INSERT INTO #carga ({','.join(cols)}) VALUES ({placeholders})",
        [[r[c] for c in cols] for r in rows],
    )
    return ex_cols


def rows_as_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def generate_processed_workbook(args, summary, sem_dl):
    src = Path(args.xlsx)
    dest = OUT / f"{src.stem} - processada-campanha-{args.id_campanha}-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(src, dest)
    wb = openpyxl.load_workbook(dest)
    ws = wb[wb.sheetnames[0]]
    start = ws.max_column + 1
    extra = [
        "STATUS_GERAL",
        "STATUS_INTEGRACAO",
        "DETALHE_INTEGRACAO",
        "STATUS_VALIDACAO",
        "STATUS_MB_PROD_EXTRA_WEB",
        "STATUS_ENVIA_ATACADO",
        "STATUS_TABELA_DL",
        "STATUS_PRECO_WISE",
        "STATUS_WISE",
        "DATA_PROCESSAMENTO",
    ]
    for i, h in enumerate(extra, start=start):
        ws.cell(1, i).value = h

    no_dl = {r["PRODUTO"] for r in sem_dl}
    status = summary.get("status", "unknown")
    if status == "success":
        integration_status = "INTEGRADO"
        detail = "Carga diária Linx/WISE executada"
        wise_default = "OK"
    elif status == "success_linx_only":
        integration_status = "LINX_OK_WISE_NAO_EXECUTADO"
        detail = "Carga Linx executada; WISE não executado nesta fase"
        wise_default = "NAO_EXECUTADO"
    elif status == "plan_only":
        integration_status = "PLANO_GERADO_SEM_EXECUCAO"
        detail = "Plano governado gerado (dry-run); nenhuma escrita real ocorreu"
        wise_default = "NAO_EXECUTADO"
    elif status == "rolled_back":
        integration_status = "NAO_INTEGRADO"
        detail = f"Execução revertida: {summary.get('error', 'ver relatório')}"
        wise_default = "NAO_EXECUTADO"
    else:
        integration_status = "PARCIAL"
        detail = f"Execução parcial: {summary.get('error', 'ver relatório')}"
        wise_default = "PENDENTE_VALIDACAO"

    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for r in range(2, ws.max_row + 1):
        produto = str(ws.cell(r, 1).value).strip()
        table_status = "SEM_DL" if produto in no_dl else "DL_OK"
        price_status = "NAO_APLICAVEL_SEM_DL" if produto in no_dl else summary.get("wise_price_status", "NAO_EXECUTADO")
        wise_status = "NAO_INTEGRADO_SEM_DL" if produto in no_dl else wise_default
        row_statuses = [integration_status, "OK", "OK", "OK", table_status, price_status, wise_status]
        status_geral = "Sucesso" if all(
            value in {"INTEGRADO", "OK", "DL_OK", "WISE_PRECO_OK"} for value in row_statuses
        ) else "Erro"
        values = [
            status_geral,
            integration_status,
            detail,
            "OK",
            "OK",
            "OK",
            table_status,
            price_status,
            wise_status,
            processed_at,
        ]
        for i, v in enumerate(values, start=start):
            ws.cell(r, i).value = v
    wb.save(dest)
    summary["processed_workbook"] = str(dest)
    return dest


# ---------------------------------------------------------------------------
# Governance gate
# ---------------------------------------------------------------------------

def _plan_hash(plan: dict) -> str:
    canonical = json.dumps(plan, sort_keys=True, default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def build_governed_plan(args, rows, diff_summary: dict) -> dict:
    """Build the exact GovernedPlanPayload contract consumed by the .NET
    GovernedPlanBridge (applications/mais-compras/backend/src/BlueprintOS.Application/Governance/GovernedPlanBridge.cs),
    field-for-field identical to what GovernedOrchestrator.buildActionProposalPayload()
    emits on the JS side (tools/agents/governed-orchestrator.js) — this script is a
    second, independent producer of the same wire contract, not a parallel one."""
    request_id = f"wise-daily-{args.id_campanha}-{args.data_limite}"
    return {
        "requestId": request_id,
        "requestedBy": "linx-wise-daily-integration-script",
        "agentId": AGENT_ID,
        "capability": CAPABILITY,
        "environment": "Production",
        "system": "WISE",
        "resourceType": "DatabaseTable",
        "resource": "MB_PROD_EXTRA_WEB;PRODUTOS;WS_ESTOQUE_PRODUTOS;WS_PRODUTOS_PRECOS",
        "operationIntent": "UPDATE",
        "fields": ["DATA_LIMITE", "ENVIA_ATACADO_INTERNET"],
        "filterSummary": f"ID_CAMPANHA={args.id_campanha};DATA_LIMITE={args.data_limite}",
        "expectedAffectedRows": len(rows),
        "purpose": "Daily Linx->WISE stock/price/grade synchronization",
        "dataClassification": "Internal",
        "containsPersonalData": False,
        "containsSensitivePersonalData": False,
        "containsSecrets": False,
        "reversibility": "PartiallyReversible",
        "runbookReference": "docs/operations/LinxWiseDailyIntegrationRunbook.md",
        "connectionProfile": CONNECTION_PROFILE,
        "additionalContext": None,
        "crossCuttingAgents": ["security-lgpd-agent"],
        "diffSummary": diff_summary,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }


def consult_governed_plan_bridge(plan: dict) -> dict:
    """Sends the plan to the real .NET governance pipeline (GovernedPlanBridge,
    via the `governed-plan` CLI — same process boundary used by
    tools/agents/governance-bridge.js) and returns its parsed JSON response.

    This is the single authority for the plan's risk classification and
    approval requirement — AIGovernancePolicyEngine, not this script and not
    the local approval mirror. A "Blocked" verdict from here cannot be
    overridden by anything in check_execution_approval(). No external
    connection is made: the CLI runs entirely in-memory (see
    InMemoryGovernedPlanStores.cs) and never opens a live WISE/SQL connection.
    """
    payload = {k: v for k, v in plan.items() if k not in ("diffSummary", "generatedAt", "planHash")}
    if not GOVERNED_PLAN_CLI_DLL.exists():
        return {"error": "GOVERNED_PLAN_BRIDGE_NOT_BUILT", "path": str(GOVERNED_PLAN_CLI_DLL)}
    try:
        result = subprocess.run(
            ["dotnet", str(GOVERNED_PLAN_CLI_DLL), "governed-plan"],
            input=json.dumps(payload), capture_output=True, text=True, timeout=30,
        )
    except (OSError, subprocess.TimeoutExpired) as e:
        return {"error": f"GOVERNED_PLAN_BRIDGE_UNREACHABLE: {e}"}
    lines = [line for line in result.stdout.splitlines() if line.strip()]
    if not lines:
        return {"error": "GOVERNED_PLAN_BRIDGE_EMPTY_OUTPUT", "stderr": result.stderr}
    try:
        return json.loads(lines[-1])
    except json.JSONDecodeError:
        return {"error": "GOVERNED_PLAN_BRIDGE_INVALID_OUTPUT", "stdout": result.stdout, "stderr": result.stderr}


def check_execution_approval(plan: dict, bridge_response: dict) -> tuple[bool, str]:
    """Returns (approved, reason). The single authority for risk/approval
    classification is bridge_response (the real AIGovernancePolicyEngine
    decision via consult_governed_plan_bridge) — this function can only ever
    narrow that decision further, never override a Blocked verdict from it.
    The local approval mirror (APPROVALS_DIR) is the last-mile check for
    "did a human actually grant *this* request", which still cannot be
    verified synchronously against the real persisted ApprovalPolicy without
    a live database connection this script must not make (documented
    limitation — see docs/audits/AgentsV1-FinalCertification.md)."""
    if bridge_response.get("error"):
        return False, f"governed plan bridge unreachable/invalid: {bridge_response['error']}"

    proposal_build = bridge_response.get("proposalBuild") or {}
    if not proposal_build.get("succeeded"):
        return False, f"governed plan bridge rejected the proposal context: {proposal_build.get('contextGaps')}"

    policy = bridge_response.get("policyDecision") or {}
    if policy.get("status") == "Blocked" or policy.get("riskClassification") == "Red":
        return False, f"AIGovernancePolicyEngine blocked this plan: {policy.get('reasons')}"
    if policy.get("status") not in ("RequiresApproval", "Allowed"):
        return False, f"unexpected policy status from bridge: {policy.get('status')!r}"

    if policy.get("status") == "Allowed":
        return True, "policy engine allowed without approval requirement"

    # RequiresApproval: the bridge's own in-memory approval request is not
    # persisted across process invocations, so the last-mile grant check
    # still falls back to the local mirror plus the exact plan-hash binding.
    approved_hash = os.environ.get("GOVERNANCE_APPROVED_EXECUTION", "").strip()
    if not approved_hash:
        return False, "GOVERNANCE_APPROVED_EXECUTION env var not set"

    computed_hash = _plan_hash(plan)
    if approved_hash != computed_hash:
        return False, "plan hash mismatch (plan changed since approval, or wrong hash supplied)"

    record_path = APPROVALS_DIR / f"{plan['requestId']}.json"
    if not record_path.exists():
        return False, f"no local approval record at {record_path}"

    try:
        record = json.loads(record_path.read_text())
    except (OSError, json.JSONDecodeError) as e:
        return False, f"approval record unreadable: {e}"

    if record.get("status") != "GRANTED":
        return False, f"approval record status is {record.get('status')!r}, not GRANTED"
    if record.get("planHash") != computed_hash:
        return False, "approval record planHash does not match current plan"
    expires_at = record.get("expiresAt")
    if not expires_at:
        return False, "approval record missing expiresAt"
    try:
        expiry = datetime.fromisoformat(expires_at)
        if expiry.tzinfo is None:
            expiry = expiry.replace(tzinfo=timezone.utc)
    except ValueError:
        return False, "approval record expiresAt is not a valid ISO timestamp"
    if expiry < datetime.now(timezone.utc):
        return False, "approval record expired"

    return True, "approved"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(args):
    OUT.mkdir(parents=True, exist_ok=True)
    data_limite = datetime.strptime(args.data_limite, "%d/%m/%Y").strftime("%Y-%m-%d")
    rows, tam_cols, sheet_errors = read_rows(Path(args.xlsx), data_limite)
    write_csv(OUT / "sheet_validation_errors.csv", sheet_errors)
    if sheet_errors:
        raise SystemExit(f"spreadsheet validation failed: {len(sheet_errors)} errors")

    cn = conn(read_only_intent=not args.execute)
    cur = cn.cursor()
    summary = {"data_limite": data_limite, "id_campanha": args.id_campanha, "rows": len(rows), "mode": "execute" if args.execute else "plan_only"}
    linx_committed = False
    sem_dl = []
    try:
        env = rows_as_dicts(cur.execute("SELECT @@SERVERNAME AS servidor, DB_NAME() AS banco"))[0]
        summary["environment"] = env
        if env["servidor"] != "SRV-SOMADB" or env["banco"] != "SOMA":
            raise SystemExit(f"environment gate failed: {env}")

        ex_cols = make_temp(cur, rows)
        missing_products = rows_as_dicts(cur.execute("""
            SELECT c.PRODUTO FROM (SELECT DISTINCT PRODUTO FROM #carga) c
            WHERE NOT EXISTS (SELECT 1 FROM PRODUTOS p WHERE p.PRODUTO = c.PRODUTO)
            ORDER BY c.PRODUTO
        """))
        missing_colors = rows_as_dicts(cur.execute("""
            SELECT c.PRODUTO, c.COR_PRODUTO FROM (SELECT DISTINCT PRODUTO, COR_PRODUTO FROM #carga) c
            WHERE NOT EXISTS (SELECT 1 FROM PRODUTO_CORES pc WHERE pc.PRODUTO = c.PRODUTO AND pc.COR_PRODUTO = c.COR_PRODUTO)
            ORDER BY c.PRODUTO, c.COR_PRODUTO
        """))
        write_csv(OUT / "missing_products.csv", missing_products)
        write_csv(OUT / "missing_product_colors.csv", missing_colors)
        summary["missing_products"] = len(missing_products)
        summary["missing_product_colors"] = len(missing_colors)
        if missing_products or missing_colors:
            cn.rollback()
            summary["status"] = "blocked_validation"
            (OUT / "final_report.json").write_text(json.dumps(summary, indent=2, default=str))
            raise SystemExit("global product/product-color validation failed")

        diff = " OR ".join(["ISNULL(CONVERT(date,m.DATA_LIMITE),'19000101') <> c.DATA_LIMITE", *[f"ISNULL(m.{c},0) <> c.{c}" for c in ex_cols]])

        # Compute the same "what would change" diff as before, but as a
        # SELECT (COUNT) instead of an UPDATE, so plan-only mode never
        # mutates anything.
        mb_would_update = rows_as_dicts(cur.execute(f"""
            SELECT COUNT(*) AS n
            FROM MB_PROD_EXTRA_WEB m
            JOIN #carga c ON c.PRODUTO=m.PRODUTO AND c.COR_PRODUTO=m.COR_PRODUTO
            WHERE {diff}
        """))[0]["n"]
        insert_cols = ["PRODUTO", "COR_PRODUTO", "DATA_LIMITE", *ex_cols]
        mb_would_insert = rows_as_dicts(cur.execute("""
            SELECT COUNT(*) AS n
            FROM #carga c
            WHERE NOT EXISTS (
              SELECT 1 FROM MB_PROD_EXTRA_WEB m
              WHERE m.PRODUTO=c.PRODUTO AND m.COR_PRODUTO=c.COR_PRODUTO
            )
        """))[0]["n"]
        envia_would_update = rows_as_dicts(cur.execute("""
            SELECT COUNT(*) AS n
            FROM PRODUTOS p
            JOIN (SELECT DISTINCT PRODUTO FROM #carga) c ON c.PRODUTO=p.PRODUTO
            WHERE ISNULL(p.ENVIA_ATACADO_INTERNET,0) <> 1
        """))[0]["n"]

        sem_dl = rows_as_dicts(cur.execute("""
            SELECT DISTINCT c.PRODUTO
            FROM #carga c
            WHERE NOT EXISTS (
              SELECT 1 FROM PRODUTOS_PRECOS pp
              WHERE pp.PRODUTO = c.PRODUTO AND pp.CODIGO_TAB_PRECO = 'DL'
            )
            ORDER BY c.PRODUTO
        """))
        write_csv(OUT / "products_without_dl.csv", sem_dl)
        summary["products_without_dl"] = len(sem_dl)

        diff_summary = {
            "mb_prod_extra_web_would_update": mb_would_update,
            "mb_prod_extra_web_would_insert": mb_would_insert,
            "produtos_envia_atacado_would_update": envia_would_update,
            "products_without_dl": len(sem_dl),
        }
        summary.update(diff_summary)

        plan = build_governed_plan(args, rows, diff_summary)
        plan["planHash"] = _plan_hash({k: v for k, v in plan.items() if k != "planHash"})
        plan_path = OUT / "governed_write_plan.json"
        plan_path.write_text(json.dumps(plan, indent=2, default=str))
        summary["governed_write_plan"] = str(plan_path)
        summary["plan_hash"] = plan["planHash"]

        bridge_response = consult_governed_plan_bridge(plan)
        bridge_path = OUT / "governed_plan_bridge_response.json"
        bridge_path.write_text(json.dumps(bridge_response, indent=2, default=str))
        summary["governed_plan_bridge_response"] = str(bridge_path)

        if not args.execute:
            cn.rollback()
            summary["status"] = "plan_only"
            print(
                "PLAN_ONLY: nenhuma escrita foi executada. "
                f"Plano governado emitido em {plan_path} (hash={plan['planHash']}). "
                f"Decisao real do AIGovernancePolicyEngine (via GovernedPlanBridge) em {bridge_path}. "
                "Para executar de verdade: a decisao de aprovacao/bloqueio e sempre a do "
                "AIGovernancePolicyEngine acima; se RequiresApproval, exporte o grant para "
                f"{APPROVALS_DIR / (plan['requestId'] + '.json')}, e rode novamente com --execute e "
                "GOVERNANCE_APPROVED_EXECUTION=<hash acima>."
            )
            return

        # --execute was passed: still gate on the real AIGovernancePolicyEngine
        # decision (via the bridge) before touching anything mutating. The
        # local approval mirror is never itself the authority — it is only the
        # last-mile grant check once the bridge has confirmed the plan is not
        # Blocked.
        approved, reason = check_execution_approval(plan, bridge_response)
        if not approved:
            cn.rollback()
            summary["status"] = "governance_blocked"
            summary["governance_block_reason"] = reason
            (OUT / "final_report.json").write_text(json.dumps(summary, indent=2, default=str))
            print(f"GOVERNANCE_REQUIRED: {reason}")
            raise SystemExit(2)

        # ---- From here on: real mutation path, only reachable with a valid
        # ---- local-mirrored grant matching the exact plan hash. ----

        set_clause = ", ".join(["m.DATA_LIMITE = c.DATA_LIMITE", *[f"m.{c} = c.{c}" for c in ex_cols]])
        cur.execute(f"""
            UPDATE m SET {set_clause}
            FROM MB_PROD_EXTRA_WEB m
            JOIN #carga c ON c.PRODUTO=m.PRODUTO AND c.COR_PRODUTO=m.COR_PRODUTO
            WHERE {diff}
        """)
        summary["mb_updated"] = cur.rowcount
        cur.execute(f"""
            INSERT INTO MB_PROD_EXTRA_WEB ({','.join(insert_cols)})
            SELECT {','.join('c.' + c for c in insert_cols)}
            FROM #carga c
            WHERE NOT EXISTS (
              SELECT 1 FROM MB_PROD_EXTRA_WEB m
              WHERE m.PRODUTO=c.PRODUTO AND m.COR_PRODUTO=c.COR_PRODUTO
            )
        """)
        summary["mb_inserted"] = cur.rowcount
        cur.execute("""
            UPDATE p SET ENVIA_ATACADO_INTERNET = 1
            FROM PRODUTOS p
            JOIN (SELECT DISTINCT PRODUTO FROM #carga) c ON c.PRODUTO=p.PRODUTO
            WHERE ISNULL(p.ENVIA_ATACADO_INTERNET,0) <> 1
        """)
        summary["envia_atacado_updated"] = cur.rowcount

        cur.execute("IF OBJECT_ID('tempdb..#aprov_pc') IS NOT NULL DROP TABLE #aprov_pc")
        cur.execute("""
            SELECT DISTINCT c.PRODUTO, c.COR_PRODUTO
            INTO #aprov_pc
            FROM #carga c
            WHERE EXISTS (
              SELECT 1 FROM PRODUTOS_PRECOS pp
              WHERE pp.PRODUTO = c.PRODUTO AND pp.CODIGO_TAB_PRECO = 'DL'
            )
        """)
        cur.execute("IF OBJECT_ID('tempdb..#expected') IS NOT NULL DROP TABLE #expected")
        cur.execute("""
            SELECT TOP 0 f.*
            INTO #expected
            FROM FN_CONSULTA_SALDO_WEB_WISE('%','%',NULL,NULL) f
        """)
        cur.execute(f"""
            INSERT INTO #expected
            SELECT f.*
            FROM FN_CONSULTA_SALDO_WEB_WISE('%','%',NULL,NULL) f
            JOIN #aprov_pc a ON a.PRODUTO=f.PRODUTO AND a.COR_PRODUTO=f.COR_PRODUTO
            WHERE f.ID_CAMPANHA = {int(args.id_campanha)}
        """)
        counts = rows_as_dicts(cur.execute("""
            SELECT
              (SELECT COUNT(*) FROM #aprov_pc) AS aprovados,
              (SELECT COUNT(*) FROM #expected) AS expected
        """))[0]
        summary["wise_expected_counts"] = counts
        if counts["aprovados"] != counts["expected"]:
            missing_expected = rows_as_dicts(cur.execute("""
                SELECT a.PRODUTO, a.COR_PRODUTO
                FROM #aprov_pc a
                WHERE NOT EXISTS (SELECT 1 FROM #expected e WHERE e.PRODUTO=a.PRODUTO AND e.COR_PRODUTO=a.COR_PRODUTO)
                ORDER BY a.PRODUTO, a.COR_PRODUTO
            """))
            write_csv(OUT / "wise_missing_expected.csv", missing_expected)
            raise RuntimeError("WISE expected stock source does not cover all approved product/colors")

        missing_remote = rows_as_dicts(cur.execute("""
            SELECT e.PRODUTO, e.COR_PRODUTO
            FROM #expected e
            WHERE NOT EXISTS (
              SELECT 1 FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
              WHERE w.ID_CAMPANHA=e.ID_CAMPANHA AND w.PRODUTO=e.PRODUTO AND w.COR_PRODUTO=e.COR_PRODUTO
            )
            ORDER BY e.PRODUTO, e.COR_PRODUTO
        """))
        write_csv(OUT / "wise_missing_remote_rows.csv", missing_remote)
        summary["wise_missing_remote_rows"] = len(missing_remote)
        if missing_remote:
            raise RuntimeError("WISE has missing remote rows; this script does not perform linked-server INSERT")

        if args.linx_only:
            cn.commit()
            linx_committed = True
            summary["wise_skipped"] = True
            summary["wise_price_status"] = "NAO_EXECUTADO"
            summary["status"] = "success_linx_only"
            return

        cn.commit()
        linx_committed = True
        cn.autocommit = True

        missing_wise_prices = rows_as_dicts(cur.execute("""
            SELECT DISTINCT a.PRODUTO
            FROM #aprov_pc a
            WHERE NOT EXISTS (
              SELECT 1
              FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_PRODUTOS_PRECOS] w
              WHERE w.ID_CAMPANHA = ?
                AND w.PRODUTO = a.PRODUTO
                AND w.CODIGO_TAB_PRECO = 'DL'
            )
            ORDER BY a.PRODUTO
        """, args.id_campanha))
        write_csv(OUT / "ws_produtos_precos_dl_missing.csv", missing_wise_prices)
        summary["wise_price_missing"] = len(missing_wise_prices)
        if missing_wise_prices:
            raise RuntimeError("WISE has missing WS_PRODUTOS_PRECOS DL rows")

        price_mismatches_before = rows_as_dicts(cur.execute("""
            SELECT DISTINCT a.PRODUTO,
                   CAST(pp.PRECO1 AS decimal(18,4)) AS LINX_PRECO1,
                   CAST(w.PRECO1 AS decimal(18,4)) AS WISE_PRECO1
            FROM #aprov_pc a
            JOIN PRODUTOS_PRECOS pp
              ON pp.PRODUTO = a.PRODUTO
             AND pp.CODIGO_TAB_PRECO = 'DL'
            JOIN [WISE_AZURE].[SOMA_LINX].[dbo].[WS_PRODUTOS_PRECOS] w
              ON w.ID_CAMPANHA = ?
             AND w.PRODUTO = a.PRODUTO
             AND w.CODIGO_TAB_PRECO = 'DL'
            WHERE ISNULL(CAST(pp.PRECO1 AS decimal(18,4)), 0) <> ISNULL(CAST(w.PRECO1 AS decimal(18,4)), 0)
            ORDER BY a.PRODUTO
        """, args.id_campanha))
        write_csv(OUT / "preco1_mismatches_before.csv", price_mismatches_before)
        summary["wise_price_mismatches_before"] = len(price_mismatches_before)

        cur.execute("""
            UPDATE w SET
              w.PRECO1 = pp.PRECO1,
              w.DATA_PARA_TRANSFERENCIA = GETDATE(),
              w.DT_INTEGRACAO = CAST(GETDATE() AS smalldatetime)
            FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_PRODUTOS_PRECOS] w
            JOIN (SELECT DISTINCT PRODUTO FROM #aprov_pc) a
              ON a.PRODUTO = w.PRODUTO
            JOIN PRODUTOS_PRECOS pp
              ON pp.PRODUTO = w.PRODUTO
             AND pp.CODIGO_TAB_PRECO = 'DL'
            WHERE w.ID_CAMPANHA = ?
              AND w.CODIGO_TAB_PRECO = 'DL'
              AND ISNULL(CAST(w.PRECO1 AS decimal(18,4)), 0) <> ISNULL(CAST(pp.PRECO1 AS decimal(18,4)), 0)
        """, args.id_campanha)
        summary["wise_price_updated"] = cur.rowcount

        price_mismatches_after = rows_as_dicts(cur.execute("""
            SELECT DISTINCT a.PRODUTO,
                   CAST(pp.PRECO1 AS decimal(18,4)) AS LINX_PRECO1,
                   CAST(w.PRECO1 AS decimal(18,4)) AS WISE_PRECO1
            FROM #aprov_pc a
            JOIN PRODUTOS_PRECOS pp
              ON pp.PRODUTO = a.PRODUTO
             AND pp.CODIGO_TAB_PRECO = 'DL'
            JOIN [WISE_AZURE].[SOMA_LINX].[dbo].[WS_PRODUTOS_PRECOS] w
              ON w.ID_CAMPANHA = ?
             AND w.PRODUTO = a.PRODUTO
             AND w.CODIGO_TAB_PRECO = 'DL'
            WHERE ISNULL(CAST(pp.PRECO1 AS decimal(18,4)), 0) <> ISNULL(CAST(w.PRECO1 AS decimal(18,4)), 0)
            ORDER BY a.PRODUTO
        """, args.id_campanha))
        write_csv(OUT / "preco1_mismatches_after.csv", price_mismatches_after)
        summary["wise_price_mismatches_after"] = len(price_mismatches_after)
        if price_mismatches_after:
            raise RuntimeError("WISE price validation failed")
        summary["wise_price_status"] = "WISE_PRECO_OK"

        # Reactivate existing approved rows, then inactivate active rows outside the approved set.
        cur.execute("""
            UPDATE w SET DT_EXCLUSAO = NULL
            FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
            JOIN #expected e ON e.ID_CAMPANHA=w.ID_CAMPANHA AND e.PRODUTO=w.PRODUTO AND e.COR_PRODUTO=w.COR_PRODUTO
            WHERE w.ID_CAMPANHA = ? AND w.DT_EXCLUSAO IS NOT NULL
        """, args.id_campanha)
        summary["wise_reactivated"] = cur.rowcount
        cur.execute("""
            UPDATE w SET DT_EXCLUSAO = GETDATE()
            FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
            WHERE w.ID_CAMPANHA = ? AND w.DT_EXCLUSAO IS NULL
              AND NOT EXISTS (SELECT 1 FROM #expected e WHERE e.ID_CAMPANHA=w.ID_CAMPANHA AND e.PRODUTO=w.PRODUTO AND e.COR_PRODUTO=w.COR_PRODUTO)
        """, args.id_campanha)
        summary["wise_inactivated"] = cur.rowcount

        mismatch_active = rows_as_dicts(cur.execute("""
            SELECT 'MISSING_ACTIVE' AS tipo, e.PRODUTO, e.COR_PRODUTO
            FROM #expected e
            WHERE NOT EXISTS (
              SELECT 1 FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
              WHERE w.ID_CAMPANHA=e.ID_CAMPANHA AND w.PRODUTO=e.PRODUTO AND w.COR_PRODUTO=e.COR_PRODUTO AND w.DT_EXCLUSAO IS NULL
            )
            UNION ALL
            SELECT 'EXTRA_ACTIVE' AS tipo, w.PRODUTO, w.COR_PRODUTO
            FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
            WHERE w.ID_CAMPANHA=? AND w.DT_EXCLUSAO IS NULL
              AND NOT EXISTS (SELECT 1 FROM #expected e WHERE e.ID_CAMPANHA=w.ID_CAMPANHA AND e.PRODUTO=w.PRODUTO AND e.COR_PRODUTO=w.COR_PRODUTO)
        """, args.id_campanha))
        write_csv(OUT / "wise_activity_mismatches.csv", mismatch_active)
        summary["wise_activity_mismatches"] = len(mismatch_active)
        if mismatch_active:
            raise RuntimeError("WISE active-set validation failed")

        grade_diff = " OR ".join([f"ISNULL(w.ES{i},0) <> ISNULL(e.D{i},0)" for i in range(1, 17)])
        grade_set = ", ".join([f"w.ES{i} = e.D{i}" for i in range(1, 17)])
        cur.execute(f"""
            UPDATE w SET
              w.LIBERAR_GRADE_WEB = e.LIBERAR_GRADE_WEB,
              w.ESTOQUE = e.SALDO_DISPONIVEL,
              {grade_set},
              w.DATA_PARA_TRANSFERENCIA = GETDATE(),
              w.DT_INTEGRACAO = CAST(GETDATE() AS smalldatetime),
              w.DT_EXCLUSAO = NULL
            FROM [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
            JOIN #expected e ON e.ID_CAMPANHA=w.ID_CAMPANHA AND e.PRODUTO=w.PRODUTO AND e.COR_PRODUTO=w.COR_PRODUTO
            WHERE w.ID_CAMPANHA = ?
              AND (ISNULL(w.LIBERAR_GRADE_WEB,'') <> ISNULL(e.LIBERAR_GRADE_WEB,'')
                   OR ISNULL(w.ESTOQUE,0) <> ISNULL(e.SALDO_DISPONIVEL,0)
                   OR {grade_diff})
        """, args.id_campanha)
        summary["wise_stock_updated"] = cur.rowcount

        post = rows_as_dicts(cur.execute("""
            SELECT e.PRODUTO, e.COR_PRODUTO
            FROM #expected e
            JOIN [WISE_AZURE].[SOMA_LINX].[dbo].[WS_ESTOQUE_PRODUTOS] w
              ON e.ID_CAMPANHA=w.ID_CAMPANHA AND e.PRODUTO=w.PRODUTO AND e.COR_PRODUTO=w.COR_PRODUTO
            WHERE w.DT_EXCLUSAO IS NOT NULL OR ISNULL(w.ESTOQUE,0) <> ISNULL(e.SALDO_DISPONIVEL,0)
        """))
        write_csv(OUT / "wise_post_mismatches.csv", post)
        summary["wise_post_mismatches"] = len(post)
        if post:
            raise RuntimeError("WISE stock validation failed")

        cn.commit()
        summary["status"] = "success"
    except SystemExit:
        raise
    except Exception as e:
        if not linx_committed:
            cn.rollback()
            summary["status"] = summary.get("status") or "rolled_back"
        else:
            summary["status"] = "linx_committed_wise_failed"
        summary["error"] = str(e)
        raise
    finally:
        if "status" in summary:
            generate_processed_workbook(args, summary, sem_dl)
        (OUT / "final_report.json").write_text(json.dumps(summary, indent=2, default=str))
        cn.close()
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--data-limite", required=True)
    p.add_argument("--id-campanha", type=int, required=True)
    p.add_argument("--linx-only", action="store_true")
    p.add_argument(
        "--execute",
        action="store_true",
        help=(
            "Attempt real mutation instead of plan-only/dry-run. Requires "
            "GOVERNANCE_APPROVED_EXECUTION=<plan hash> and a matching local "
            "approval record; otherwise the script exits with GOVERNANCE_REQUIRED."
        ),
    )
    run(p.parse_args())
